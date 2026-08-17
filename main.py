"""
main.py
-------
Streamlit entry point for the Excel processing web application.

Structure:
  - Sidebar: action buttons for each available report
  - Main area: dynamic content based on the selected action
"""

import datetime
import io

import streamlit as st

from utils.excel_utils import load_sheets, dfs_to_excel_bytes
from processors import internet_morchav

APP_VERSION_UPDATED_AT = "17.08.2026 12:01"

FILTERED_ORIGINAL_SELLERS = ("אליאור ביטון", "זהבה בלאי", "עומר בר מוחה")
SELLER_FILTER_COLUMNS = {
    "סיבים": 7,      # Column G: שם מוכרן
    "נחושת": 7,      # Column G: שם מוכרן
    "כל השאר": 9,   # Column I
}


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _uploaded_file_bytes(uploaded_file) -> bytes:
    if isinstance(uploaded_file, bytes):
        return uploaded_file
    return uploaded_file.getvalue()


def _filter_worksheet_by_seller(worksheet, seller_col_idx: int, allowed_sellers: set[str]) -> None:
    rows_to_delete = []
    for row_idx in range(worksheet.max_row, 2, -1):
        seller_name = _cell_text(worksheet.cell(row=row_idx, column=seller_col_idx).value)
        if seller_name not in allowed_sellers:
            rows_to_delete.append(row_idx)

    run_start = None
    previous_row = None
    for row_idx in rows_to_delete:
        if run_start is None:
            run_start = previous_row = row_idx
            continue

        if row_idx == previous_row - 1:
            previous_row = row_idx
            continue

        worksheet.delete_rows(previous_row, run_start - previous_row + 1)
        run_start = previous_row = row_idx

    if run_start is not None:
        worksheet.delete_rows(previous_row, run_start - previous_row + 1)


def filtered_original_workbook_bytes(uploaded_file) -> bytes:
    """
    Return a copy of the original workbook with only selected seller rows.
    Rows 1-2 are kept; filtering starts from row 3.
    """
    from openpyxl import load_workbook

    source_bytes = _uploaded_file_bytes(uploaded_file)
    workbook = load_workbook(io.BytesIO(source_bytes))
    allowed_sellers = {_cell_text(seller) for seller in FILTERED_ORIGINAL_SELLERS}

    for sheet_name, seller_col_idx in SELLER_FILTER_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"חסר גיליון נדרש עבור דוח מוכרנים: {sheet_name}")

        worksheet = workbook[sheet_name]
        _filter_worksheet_by_seller(worksheet, seller_col_idx, allowed_sellers)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="מערכת דוחות אקסל",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── RTL styling ───────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
        body, .stApp { direction: rtl; text-align: right; }
        .stButton button,
        .stDownloadButton button {
            width: 100%;
            min-height: 2.4rem;
            border-radius: 0.45rem;
            font-weight: 600;
            white-space: normal;
            line-height: 1.2;
            padding: 0.45rem 0.75rem;
        }
        .stDownloadButton button {
            background-color: #0a7c59;
            color: white;
        }
        .stDownloadButton button:disabled {
            background-color: #e8edf3;
            color: #8c98a8;
        }
        .block-container { padding-top: 2rem; }
        div[data-testid="stFileUploader"] {
            margin-bottom: 0.15rem;
        }
        div[data-testid="stFileUploader"] section {
            min-height: 3.6rem;
            padding: 0.55rem 0.75rem;
            border-radius: 0.5rem;
        }
        div[data-testid="stHorizontalBlock"] {
            gap: 0.65rem;
        }
        .action-offset {
            height: 1.8rem;
        }
        .status-summary {
            display: flex;
            width: fit-content;
            max-width: 100%;
            flex-wrap: wrap;
            align-items: center;
            gap: 0.6rem;
            margin: 0.9rem 0 0.65rem 0;
            padding: 0.75rem 0.9rem;
            border: 1px solid #b7ebcc;
            border-radius: 0.5rem;
            background: #ecfdf3;
            color: #087443;
            font-weight: 600;
        }
        .status-summary span {
            display: inline-flex;
            align-items: center;
            padding: 0.18rem 0.55rem;
            border-radius: 999px;
            background: rgba(255, 255, 255, 0.72);
            border: 1px solid rgba(10, 124, 89, 0.14);
        }
        .download-heading {
            margin: 0.45rem 0 0.35rem 0;
            font-size: 1rem;
            font-weight: 700;
            color: #243044;
        }
        .soft-note {
            width: fit-content;
            max-width: 100%;
            margin: 0.5rem 0 0.2rem 0;
            padding: 0.45rem 0.65rem;
            border: 1px solid #b7ebcc;
            border-radius: 0.45rem;
            background: #ecfdf3;
            color: #0a7c59;
            font-size: 0.9rem;
            font-weight: 600;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Sidebar – action selection ────────────────────────────────────────────────
st.sidebar.title("📂 פעולות")
st.sidebar.markdown("---")

ACTIONS = {
    "internet_morchav": "🌐 ניתוח סטטוס אינטרנט מחודש להרצה",
}

# Keep the selected action in session state so clicking doesn't reset the page
if "selected_action" not in st.session_state:
    st.session_state.selected_action = None

for key, label in ACTIONS.items():
    if st.sidebar.button(label, key=f"btn_{key}"):
        st.session_state.selected_action = key
        # Clear any previous results when switching actions
        st.session_state.pop("analysis_result", None)

# ── Main area ─────────────────────────────────────────────────────────────────
st.title("📊 מערכת דוחות אקסל")
st.caption(f"גרסא: {APP_VERSION_UPDATED_AT}")

if st.session_state.selected_action is None:
    st.info("בחר פעולה מהתפריט משמאל כדי להתחיל.")
    st.stop()

# ── Action: ניתוח סטטוס אינטרנט מורכב להרצה ──────────────────────────────────
if st.session_state.selected_action == "internet_morchav":
    st.header("🌐 ניתוח סטטוס אינטרנט מחודש להרצה")
    st.markdown(
        "העלה את קובץ ה-Excel המכיל את הגיליונות **סיבים**, **נחושת** ו-**כל השאר**."
    )

    run_analysis = False
    prepare_mg_report = False
    upload_col, action_col, mg_col, mg_download_col = st.columns([5.3, 1.15, 1.15, 1.1])

    with upload_col:
        uploaded = st.file_uploader(
            "בחר קובץ Excel",
            type=["xlsx"],
            key="upload_internet_morchav",
        )

    if uploaded:
        uploaded_signature = (uploaded.name, uploaded.size)
        if st.session_state.get("uploaded_file_signature") != uploaded_signature:
            st.session_state["uploaded_file_signature"] = uploaded_signature
            st.session_state.pop("analysis_result", None)
            st.session_state.pop("mg_report_bytes", None)
    else:
        st.session_state.pop("uploaded_file_signature", None)
        st.session_state.pop("analysis_result", None)
        st.session_state.pop("mg_report_bytes", None)

    with action_col:
        st.markdown('<div class="action-offset"></div>', unsafe_allow_html=True)
        run_analysis = st.button(
            "▶️ הפעל ניתוח",
            key="run_internet_morchav",
            disabled=uploaded is None,
        )

    with mg_col:
        st.markdown('<div class="action-offset"></div>', unsafe_allow_html=True)
        prepare_mg_report = st.button(
            "⚙️ דוח עבור MG",
            key="prepare_mg_report_top",
            disabled=uploaded is None,
        )

    with mg_download_col:
        st.markdown('<div class="action-offset"></div>', unsafe_allow_html=True)
        if uploaded and "mg_report_bytes" in st.session_state:
            st.download_button(
                label="⬇️ הורד MG",
                data=st.session_state["mg_report_bytes"],
                file_name=f"דוח עבור MG - {datetime.date.today().strftime('%d.%m.%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_mg_report_top",
            )

    if uploaded:
        if prepare_mg_report:
            with st.spinner("מכין דוח עבור MG..."):
                try:
                    filtered_original_bytes = filtered_original_workbook_bytes(uploaded)
                    st.session_state["mg_report_bytes"] = filtered_original_bytes
                    st.markdown('<div class="soft-note">✅ דוח עבור MG מוכן להורדה.</div>', unsafe_allow_html=True)
                    mg_download_col.download_button(
                        label="⬇️ הורד MG",
                        data=filtered_original_bytes,
                        file_name=f"דוח עבור MG - {datetime.date.today().strftime('%d.%m.%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_mg_report_top_ready",
                    )
                except Exception as e:
                    import traceback
                    st.error("❌ שגיאה בהכנת דוח עבור MG")
                    st.markdown(
                        f"""
**סוג השגיאה:** `{type(e).__name__}`

**פירוט:** `{e}`

**מה לבדוק:**
- האם שמות הגיליונות בקובץ הם בדיוק: `סיבים`, `נחושת`, `כל השאר`?
- האם קיימות עמודות הפילטר הנדרשות: G בגיליונות `סיבים`/`נחושת`, ו-I בגיליון `כל השאר`?
"""
                    )
                    with st.expander("🔍 פרטי שגיאה מלאים (Traceback)"):
                        st.code(traceback.format_exc(), language="python")

        # Run analysis button
        if run_analysis:
            with st.spinner("מנתח את הקובץ..."):
                try:
                    sheets = load_sheets(
                        uploaded,
                        sheet_names=["סיבים", "נחושת", "כל השאר"],
                    )

                    analysis_output = internet_morchav.run(
                        fiber_df  = sheets["סיבים"],
                        copper_df = sheets["נחושת"],
                        rest_df   = sheets["כל השאר"],
                    )
                    if len(analysis_output) == 4:
                        result_df, exceptions_df, phone_df, biznet_df = analysis_output
                    else:
                        result_df, exceptions_df, phone_df = analysis_output
                        biznet_df = result_df.iloc[0:0].copy()

                    st.session_state["analysis_result"] = {
                        "result":     result_df,
                        "exceptions": exceptions_df,
                        "phone":      phone_df,
                        "biznet":     biznet_df,
                        "source_workbook": uploaded.getvalue(),
                    }
                except Exception as e:
                    import traceback
                    st.error("❌ שגיאה בעיבוד הקובץ")
                    st.markdown(
                        f"""
**סוג השגיאה:** `{type(e).__name__}`

**פירוט:** `{e}`

**מה לבדוק:**
- האם שמות הגיליונות בקובץ הם בדיוק: `סיבים`, `נחושת`, `כל השאר`?
- האם שורת הכותרות נמצאת בשורה **2** של הגיליון?
- האם כל העמודות הנדרשות קיימות בגיליון?
"""
                    )
                    with st.expander("🔍 פרטי שגיאה מלאים (Traceback)"):
                        st.code(traceback.format_exc(), language="python")
                    st.stop()

    # Display results if available
    if "analysis_result" in st.session_state:
        data = st.session_state["analysis_result"]
        result_df     = data["result"]
        exceptions_df = data["exceptions"]
        phone_df      = data["phone"]
        biznet_df     = data.get("biznet", result_df.iloc[0:0].copy())
        biznet_df     = biznet_df.drop(columns=["תאריך ושעת התקנה מעודכנים"], errors="ignore")

        # ── Split result by "תאריך מתואם" ─────────────────────────────────
        coord_col = "תאריך מתואם"
        has_date_mask = (
            result_df[coord_col].notna()
            & (result_df[coord_col].astype(str).str.strip() != "")
            & (result_df[coord_col].astype(str).str.strip().str.lower() != "nan")
        )
        result_with_date    = result_df[has_date_mask].reset_index(drop=True)
        result_without_date = result_df[~has_date_mask].drop(columns=[coord_col]).reset_index(drop=True)
        st.markdown(
            f"""
<div class="status-summary">
    <strong>✅ הניתוח הושלם</strong>
    <span>עם תאריך: {len(result_with_date)}</span>
    <span>ללא תאריך: {len(result_without_date)}</span>
    <span>BIZNET: {len(biznet_df)}</span>
    <span>קו טלפון: {len(phone_df)}</span>
</div>
""",
            unsafe_allow_html=True,
        )

        today_str = datetime.date.today().strftime("%d.%m.%Y")

        sheets_with_date = {"סטטוס הזמנות": result_with_date}
        if not exceptions_df.empty:
            sheets_with_date["חריגים"] = exceptions_df

        sheets_without_date = {"סטטוס הזמנות": result_without_date}
        if not exceptions_df.empty:
            sheets_without_date["חריגים"] = exceptions_df

        st.markdown('<div class="download-heading">הורדות</div>', unsafe_allow_html=True)
        download_cols = st.columns([1.35, 1.35, 1.05, 1.05, 3.8])
        with download_cols[0]:
            st.download_button(
                label="⬇️ אינטרנט - עם תאריך",
                data=dfs_to_excel_bytes(sheets_with_date),
                file_name=f"סטטוס אינטרנט מורכב להרצה - עם תאריך - {today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_internet_with_date_top",
            )

        with download_cols[1]:
            st.download_button(
                label="⬇️ אינטרנט - ללא תאריך",
                data=dfs_to_excel_bytes(sheets_without_date),
                file_name=f"סטטוס אינטרנט מורכב להרצה - ללא תאריך - {today_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_internet_without_date_top",
            )

        with download_cols[2]:
            if not biznet_df.empty:
                st.download_button(
                    label="⬇️ BIZNET",
                    data=dfs_to_excel_bytes({"הזמנות BIZNET": biznet_df}),
                    file_name=f"סטטוס BIZNET - {today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_biznet_top",
                )

        with download_cols[3]:
            if not phone_df.empty:
                st.download_button(
                    label="⬇️ קו טלפון",
                    data=dfs_to_excel_bytes({"הזמנות קו טלפון": phone_df}),
                    file_name=f"סטטוס קו טלפון - {today_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_phone_top",
                )

        # ── Preview: with date ─────────────────────────────────────────────
        st.subheader(f"📋 סטטוס אינטרנט – עם תאריך מתואם ({len(result_with_date)} שורות)")
        st.dataframe(result_with_date, use_container_width=True)

        # ── Preview: without date ──────────────────────────────────────────
        st.subheader(f"📋 סטטוס אינטרנט – ללא תאריך מתואם ({len(result_without_date)} שורות)")
        st.dataframe(result_without_date, use_container_width=True)

        if not exceptions_df.empty:
            st.warning(f"⚠️ נמצאו {len(exceptions_df)} שורות חריגות (סטטוס שירות לא מוכר).")
            with st.expander("הצג חריגים"):
                st.dataframe(exceptions_df, use_container_width=True)

        # ── Phone result preview ───────────────────────────────────────────
        if not phone_df.empty:
            st.subheader("📞 תצוגה מקדימה – הזמנות קו טלפון")
            st.dataframe(phone_df, use_container_width=True)

        # ── BIZNET result preview ─────────────────────────────────────────
        if not biznet_df.empty:
            st.subheader("🌐 תצוגה מקדימה – הזמנות BIZNET")
            st.dataframe(biznet_df, use_container_width=True)
