"""
Compare an original status workbook with an exported MG workbook.

Checks:
1. Sheet names, columns, kept row values, cell data types, number formats, and style ids.
2. XLSX package structure differences at the ZIP/XML level.

Usage:
    python tools/compare_mg_workbooks.py "original.xlsx" "mg.xlsx"
"""

from __future__ import annotations

import argparse
import hashlib
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


FILTERED_ORIGINAL_SELLERS = (
    "\u05d0\u05dc\u05d9\u05d0\u05d5\u05e8 \u05d1\u05d9\u05d8\u05d5\u05df",
    "\u05d6\u05d4\u05d1\u05d4 \u05d1\u05dc\u05d0\u05d9",
    "\u05e2\u05d5\u05de\u05e8 \u05d1\u05e8 \u05de\u05d5\u05d7\u05d4",
)
SELLER_FILTER_COLUMNS = {
    "\u05e1\u05d9\u05d1\u05d9\u05dd": 7,
    "\u05e0\u05d7\u05d5\u05e9\u05ea": 7,
    "\u05db\u05dc \u05d4\u05e9\u05d0\u05e8": 9,
}


@dataclass
class Issue:
    level: str
    message: str


def cell_signature(cell) -> tuple:
    return (
        cell.value,
        cell.data_type,
        cell.number_format,
        cell.style_id,
        cell.is_date,
    )


def cell_label(row: int, col: int) -> str:
    return f"R{row}C{col}"


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def iter_used_cells(ws) -> Iterable[tuple[int, int]]:
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            yield row, col


def compare_sheet_names(source_wb, mg_wb, issues: list[Issue]) -> None:
    if source_wb.sheetnames != mg_wb.sheetnames:
        issues.append(
            Issue(
                "FAIL",
                f"Sheet names/order changed: source={source_wb.sheetnames}, mg={mg_wb.sheetnames}",
            )
        )


def expected_rows_after_filter(source_ws, seller_col_idx: int) -> list[int]:
    allowed = {text(seller) for seller in FILTERED_ORIGINAL_SELLERS}
    rows = [1, 2]
    for row_idx in range(3, source_ws.max_row + 1):
        if text(source_ws.cell(row=row_idx, column=seller_col_idx).value) in allowed:
            rows.append(row_idx)
    return rows


def compare_cell(source_cell, mg_cell, sheet_name: str, source_row: int, mg_row: int, col: int, issues: list[Issue]) -> None:
    if cell_signature(source_cell) != cell_signature(mg_cell):
        issues.append(
            Issue(
                "FAIL",
                (
                    f"{sheet_name} {cell_label(mg_row, col)} differs from source {cell_label(source_row, col)}: "
                    f"source=(value={source_cell.value!r}, type={source_cell.data_type!r}, "
                    f"format={source_cell.number_format!r}, style={source_cell.style_id}, is_date={source_cell.is_date}) "
                    f"mg=(value={mg_cell.value!r}, type={mg_cell.data_type!r}, "
                    f"format={mg_cell.number_format!r}, style={mg_cell.style_id}, is_date={mg_cell.is_date})"
                ),
            )
        )


def compare_filtered_sheet(source_ws, mg_ws, sheet_name: str, seller_col_idx: int, issues: list[Issue]) -> None:
    expected_rows = expected_rows_after_filter(source_ws, seller_col_idx)
    if mg_ws.max_column != source_ws.max_column:
        issues.append(
            Issue(
                "FAIL",
                f"{sheet_name}: column count changed: source={source_ws.max_column}, mg={mg_ws.max_column}",
            )
        )
    if mg_ws.max_row != len(expected_rows):
        issues.append(
            Issue(
                "FAIL",
                f"{sheet_name}: row count mismatch after filter: expected={len(expected_rows)}, mg={mg_ws.max_row}",
            )
        )

    rows_to_compare = min(mg_ws.max_row, len(expected_rows))
    cols_to_compare = min(mg_ws.max_column, source_ws.max_column)
    for mg_row in range(1, rows_to_compare + 1):
        source_row = expected_rows[mg_row - 1]
        for col in range(1, cols_to_compare + 1):
            compare_cell(
                source_ws.cell(row=source_row, column=col),
                mg_ws.cell(row=mg_row, column=col),
                sheet_name,
                source_row,
                mg_row,
                col,
                issues,
            )


def compare_unfiltered_sheet(source_ws, mg_ws, sheet_name: str, issues: list[Issue]) -> None:
    if (source_ws.max_row, source_ws.max_column) != (mg_ws.max_row, mg_ws.max_column):
        issues.append(
            Issue(
                "FAIL",
                (
                    f"{sheet_name}: dimensions changed: "
                    f"source={source_ws.max_row}x{source_ws.max_column}, mg={mg_ws.max_row}x{mg_ws.max_column}"
                ),
            )
        )

    rows_to_compare = min(source_ws.max_row, mg_ws.max_row)
    cols_to_compare = min(source_ws.max_column, mg_ws.max_column)
    for row in range(1, rows_to_compare + 1):
        for col in range(1, cols_to_compare + 1):
            compare_cell(
                source_ws.cell(row=row, column=col),
                mg_ws.cell(row=row, column=col),
                sheet_name,
                row,
                row,
                col,
                issues,
            )


def compare_workbook_content(source_path: Path, mg_path: Path, issues: list[Issue]) -> None:
    source_wb = load_workbook(source_path, data_only=False)
    mg_wb = load_workbook(mg_path, data_only=False)

    compare_sheet_names(source_wb, mg_wb, issues)

    for sheet_name in source_wb.sheetnames:
        if sheet_name not in mg_wb.sheetnames:
            continue

        source_ws = source_wb[sheet_name]
        mg_ws = mg_wb[sheet_name]
        if sheet_name in SELLER_FILTER_COLUMNS:
            compare_filtered_sheet(source_ws, mg_ws, sheet_name, SELLER_FILTER_COLUMNS[sheet_name], issues)
        else:
            compare_unfiltered_sheet(source_ws, mg_ws, sheet_name, issues)


def zip_entry_hashes(path: Path) -> dict[str, str]:
    hashes = {}
    with zipfile.ZipFile(path, "r") as archive:
        for name in sorted(archive.namelist()):
            hashes[name] = hashlib.sha256(archive.read(name)).hexdigest()
    return hashes


def compare_xlsx_package(source_path: Path, mg_path: Path, issues: list[Issue]) -> None:
    source_hashes = zip_entry_hashes(source_path)
    mg_hashes = zip_entry_hashes(mg_path)

    source_entries = set(source_hashes)
    mg_entries = set(mg_hashes)
    added = sorted(mg_entries - source_entries)
    removed = sorted(source_entries - mg_entries)
    changed = sorted(name for name in source_entries & mg_entries if source_hashes[name] != mg_hashes[name])

    if added:
        issues.append(Issue("WARN", f"XLSX package entries added: {added}"))
    if removed:
        issues.append(Issue("WARN", f"XLSX package entries removed: {removed}"))
    if changed:
        preview = changed[:30]
        suffix = "" if len(changed) <= 30 else f" ... and {len(changed) - 30} more"
        issues.append(Issue("WARN", f"XLSX package entries changed ({len(changed)}): {preview}{suffix}"))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path, help="Original source workbook")
    parser.add_argument("mg", type=Path, help="Generated MG workbook")
    args = parser.parse_args()

    issues: list[Issue] = []
    compare_workbook_content(args.source, args.mg, issues)
    compare_xlsx_package(args.source, args.mg, issues)

    fails = [issue for issue in issues if issue.level == "FAIL"]
    warnings = [issue for issue in issues if issue.level == "WARN"]

    print("=== MG workbook comparison ===")
    if not fails:
        print("PASS: Columns, kept values, data types, number formats, and style ids match the source rows.")
    else:
        print(f"FAIL: Found {len(fails)} content/format mismatch(es).")
    if warnings:
        print(f"WARN: Found {len(warnings)} XLSX package-level difference group(s).")
    else:
        print("PASS: XLSX package entries are byte-identical.")

    for issue in issues[:200]:
        print(f"{issue.level}: {issue.message}")
    if len(issues) > 200:
        print(f"... {len(issues) - 200} additional issues omitted")

    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
