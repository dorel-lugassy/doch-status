"""
excel_utils.py
--------------
Shared helpers for loading and saving Excel files.
"""

import io
from copy import copy
from typing import Optional, Set

import pandas as pd
from openpyxl import load_workbook


BIZNET_SHEET_NAME = "הזמנות BIZNET"
COORD_DATE_COL = "תאריך מתואם"
FILTERED_ORIGINAL_SELLERS = ("אליאור ביטון", "זהבה בלאי", "עומר בר מוחה")
SELLER_FILTER_COLUMNS = {
    "סיבים": 7,      # Column G: שם מוכרן
    "נחושת": 7,      # Column G: שם מוכרן
    "כל השאר": 9,   # Column I
}


def load_sheets(uploaded_file, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
    """
    Load specific sheets from an uploaded Streamlit file object.

    Parameters
    ----------
    uploaded_file : UploadedFile
        The file object from st.file_uploader.
    sheet_names : list[str]
        List of sheet names to load.

    Returns
    -------
    dict[str, pd.DataFrame]
        Mapping of sheet name → DataFrame (columns stripped of whitespace).
    """
    result = {}
    for name in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=name, dtype=str, header=1)
        # Strip surrounding whitespace from column names and string values
        df.columns = [str(c).strip() for c in df.columns]
        df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
        # Drop repeated header rows (the file sometimes has a header row at the bottom)
        order_col = "מספר הזמנה"
        if order_col in df.columns:
            df = df[df[order_col] != order_col].reset_index(drop=True)
        result[name] = df
    return result


def dfs_to_excel_bytes(
    sheets: dict[str, pd.DataFrame],
    text_columns: Optional[Set[str]] = None,
) -> bytes:
    """
    Serialize one or more DataFrames into an in-memory Excel file.

    Parameters
    ----------
    sheets : dict[str, pd.DataFrame]
        Mapping of sheet name → DataFrame to write.
    text_columns : set[str] | None
        Column names to force as Excel text cells.

    Returns
    -------
    bytes
        Raw bytes of the .xlsx file, ready for st.download_button.
    """
    text_columns = text_columns or set()
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            sheet_text_columns = set(text_columns)
            if sheet_name == BIZNET_SHEET_NAME:
                sheet_text_columns.add(COORD_DATE_COL)
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name not in sheet_text_columns:
                    continue
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)
    return buffer.getvalue()


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _copy_row(worksheet, source_row_idx: int, target_row_idx: int) -> None:
    if source_row_idx == target_row_idx:
        return

    source_dim = worksheet.row_dimensions[source_row_idx]
    target_dim = worksheet.row_dimensions[target_row_idx]
    target_dim.height = source_dim.height
    target_dim.hidden = source_dim.hidden
    target_dim.outlineLevel = source_dim.outlineLevel
    target_dim.collapsed = source_dim.collapsed

    for col_idx in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(row=source_row_idx, column=col_idx)
        target_cell = worksheet.cell(row=target_row_idx, column=col_idx)

        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.comment = copy(source_cell.comment)
        target_cell.hyperlink = copy(source_cell.hyperlink)


def _filter_worksheet_by_seller(worksheet, seller_col_idx: int, allowed_sellers: set[str]) -> None:
    keep_rows = [1, 2]
    for row_idx in range(3, worksheet.max_row + 1):
        seller_name = _cell_text(worksheet.cell(row=row_idx, column=seller_col_idx).value)
        if seller_name in allowed_sellers:
            keep_rows.append(row_idx)

    original_max_row = worksheet.max_row
    for target_row_idx, source_row_idx in enumerate(keep_rows, start=1):
        _copy_row(worksheet, source_row_idx, target_row_idx)

    first_delete_row = len(keep_rows) + 1
    if first_delete_row <= original_max_row:
        worksheet.delete_rows(first_delete_row, original_max_row - first_delete_row + 1)


def filtered_original_workbook_bytes(uploaded_file) -> bytes:
    """
    Return a copy of the original workbook with only selected seller rows.

    The workbook is edited directly with openpyxl so original sheet names,
    columns, values, spacing, and formatting are preserved as much as possible.
    Rows 1-2 are kept; filtering starts from row 3.
    """
    source_bytes = uploaded_file.getvalue()
    workbook = load_workbook(io.BytesIO(source_bytes))
    allowed_sellers = {_cell_text(seller) for seller in FILTERED_ORIGINAL_SELLERS}

    for sheet_name, seller_col_idx in SELLER_FILTER_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"חסר גיליון נדרש עבור דוח מוכרנים: {sheet_name}")

        worksheet = workbook[sheet_name]
        _filter_worksheet_by_seller(worksheet, seller_col_idx, allowed_sellers)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
